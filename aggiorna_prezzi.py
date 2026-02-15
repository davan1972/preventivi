import openpyxl
import json
import re

# File paths
FILE_POLTRONCINE = 'LISTINI_2026_FACILE_SALIRE_-_MS_POLTRONCINE.xlsx'
FILE_PEDANE = 'LISTINI_2026_FACILE_SALIRE_-_MS_PEDANE.xlsx'
HTML_FILE = 'preventivi-montascale-v2.html'

print("🔄 Estrazione prezzi da Excel...")

# Funzione per estrarre prezzi da un foglio Excel
def estrai_prezzi_foglio(ws):
    prezzi = {}
    
    # Prezzo base (riga 6, colonna D)
    if ws.cell(6, 4).value:
        prezzi['base'] = float(ws.cell(6, 4).value)
    
    # Scansiona tutte le righe per trovare opzioni
    for row_idx in range(7, min(ws.max_row + 1, 60)):
        descrizione = ws.cell(row_idx, 2).value  # Colonna B
        prezzo = ws.cell(row_idx, 4).value  # Colonna D
        
        if descrizione and prezzo and isinstance(prezzo, (int, float)):
            desc_lower = str(descrizione).lower()
            
            # Mappa descrizioni a ID opzioni
            mappings = {
                'estensione guida': 'opt_estensione_guida',
                'guida tagliata': 'opt_guida_tagliata',
                'guida ribaltabile manuale': 'opt_guida_ribaltabile',
                'tappeti': 'opt_tappeti',
                'guida scorrevole 4.7': 'opt_guida_scorrevole_47',
                'guida scorrevole 7.05': 'opt_guida_scorrevole_705',
                'zero intrusion 4.7': 'opt_zero_intrusion_47',
                'zero intrusion 7.05': 'opt_zero_intrusion_705',
                'rotazione del seggiolino': 'opt_rotazione_auto',
                'rotazione automatica sedile 2': 'opt_rotazione_2dir',
                'poggiapiedi motorizzato': 'opt_pedana_motor',
                'motorizzato completo (sedile': 'opt_motor_completo',
                'motorizzato completo 2': 'opt_motor_completo_2dir',
                'one touch': 'opt_one_touch',
                'sedile smart perch': 'opt_sedile_perch',
                'sedile smart (versione': 'opt_sedile_esterni',
                'sedile style': 'opt_sedile_style',
                'bracciolo corto': 'opt_bracciolo_corto',
                'distanziale per bracciolo singolo': 'opt_distanziale_singolo',
                'distanziale entrambi': 'opt_distanziale_doppio',
                'cintura in velcro': 'opt_cintura_velcro',
                'body harness': 'opt_body_harness',
                'heavy duty xl': 'opt_heavy_duty_xl',
                'extra heavy duty': 'opt_heavy_duty_xxl',
                'versione per esterni': 'opt_versione_esterni',
                'guida ribaltabile motorizzata': 'opt_guida_ribaltabile_motor',
                'smontaggio e smaltimento': 'opt_smontaggio',
                'overrun': 'opt_overrun',
                'dropnose': 'opt_dropnose',
                'guida al mt': 'metri_guida',
                'curva aggiuntiva': 'curve_90_qty',
                'punti di ricarica': 'punti_ricarica_qty',
                'guida per esterni per mt': 'metri_esterni',
                'fermata intermedia': 'fermate_intermedie_qty',
                'radiocomando aggiuntivo': 'radiocomandi_qty',
                'ogni alzata': 'alzate',
                'ogni metro di guida': 'metri_guida_extra',
                'colonnino chiamate': 'colonnini_qty',
                'elegance seat': 'opt_elegance_seat',
                'alliance seat': 'opt_alliance_seat',
            }
            
            for keyword, opt_id in mappings.items():
                if keyword in desc_lower:
                    prezzi[opt_id] = float(prezzo)
                    break
    
    return prezzi

# Estrai prezzi da tutti i modelli
modelli_prezzi = {}

print("📊 Elaborazione file Poltroncine...")
wb_poltroncine = openpyxl.load_workbook(FILE_POLTRONCINE, data_only=True)
for sheet_name in wb_poltroncine.sheetnames:
    print(f"  ✓ {sheet_name}")
    ws = wb_poltroncine[sheet_name]
    modelli_prezzi[sheet_name] = estrai_prezzi_foglio(ws)

print("📊 Elaborazione file Pedane...")
wb_pedane = openpyxl.load_workbook(FILE_PEDANE, data_only=True)
for sheet_name in wb_pedane.sheetnames:
    print(f"  ✓ {sheet_name}")
    ws = wb_pedane[sheet_name]
    modelli_prezzi[sheet_name] = estrai_prezzi_foglio(ws)

# Genera codice JavaScript
js_code = "const prezziModelli = " + json.dumps(modelli_prezzi, indent=12) + ";"

# Leggi HTML
print("\n📝 Aggiornamento HTML...")
with open(HTML_FILE, 'r', encoding='utf-8') as f:
    html_content = f.read()

# Trova e sostituisci sezione prezziModelli
pattern = r'const prezziModelli = \{[\s\S]*?\};'
new_html = re.sub(pattern, js_code, html_content)

# Salva HTML aggiornato
with open(HTML_FILE, 'w', encoding='utf-8') as f:
    f.write(new_html)

print("✅ HTML aggiornato con successo!")
print(f"\n📦 Modelli elaborati: {len(modelli_prezzi)}")
for modello, prezzi in modelli_prezzi.items():
    print(f"  • {modello}: {len(prezzi)} voci")

print("\n🎯 Prossimi passi:")
print("1. Verifica il file 'preventivi-montascale-v2.html'")
print("2. Carica su GitHub")
print("3. Fatto!")
